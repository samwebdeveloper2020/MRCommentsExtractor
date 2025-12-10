"""
Main GUI Window for GitLab MR Comments Viewer
Provides a user-friendly interface to extract and view merge request comments
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import threading
import json
import os
from services.gitlab_api import GitLabAPI
from services.llm_service import LLMService
from utils.helpers import parse_gitlab_url, format_datetime, count_comments, extract_comment_text, get_file_info_from_position, extract_images_from_text, replace_images_in_text, get_code_context_from_discussion
from utils.token_manager import TokenManager
from utils.image_viewer import ImageViewer

class MainWindow:
    def __init__(self, root):
        """Initialize the main window
        
        Args:
            root: Tkinter root window
        """
        print("DEBUG: MainWindow __init__ called")
        self.root = root
        self.root.title("GitLab MR Comments Viewer - Code Review Assistant")
        self.root.geometry("1200x800")
        
        # Configure modern styling
        self.configure_styles()
        
        # Variables
        self.url_var = tk.StringVar()
        self.project_var = tk.StringVar()
        self.mr_var = tk.StringVar()
        self.mr_state_var = tk.StringVar()
        self.mr_created_var = tk.StringVar()
        self.mr_merged_var = tk.StringVar()
        self.mr_assignees_var = tk.StringVar()
        self.comments_data = None
        self.downloaded_images = {}
        self.projects_data = []
        self.all_project_names = []  # Store all project names for filtering
        self.current_mrs = []
        self.all_mr_names = []  # Store all MR names for filtering
        self.is_filtering = False  # Flag to prevent recursive filtering
        self.is_filtering_mrs = False  # Flag to prevent recursive MR filtering
        
        # Initialize token manager and image viewer
        self.token_manager = TokenManager()
        self.image_viewer = ImageViewer(root)
        
        # Load tokens from files BEFORE setup_ui
        token, _, _ = self.token_manager.load_token()
        self.gitlab_token = token if token else None
        
        llm_token = self.token_manager.load_llm_token()
        self.llm_token = llm_token if llm_token else None
        
        self.setup_ui()
        
        # Update status based on token availability
        if self.gitlab_token and self.llm_token:
            self.status_var.set("Ready - Tokens loaded from files")
        elif self.gitlab_token:
            self.status_var.set("Ready - GitLab token loaded (LLM token missing)")
        elif self.llm_token:
            self.status_var.set("Ready - LLM token loaded (GitLab token missing)")
        else:
            self.status_var.set("Ready - No tokens found. Please add tokens to token.json and llm_token.json")
    
    def configure_styles(self):
        """Configure modern UI styles"""
        style = ttk.Style()
        
        # Try to use a modern theme
        try:
            style.theme_use('clam')  # Modern, clean theme
        except:
            pass
        
        # Configure colors
        bg_color = "#f5f5f5"
        fg_color = "#2c3e50"
        accent_color = "#3498db"
        
        # Configure TFrame
        style.configure('TFrame', background=bg_color)
        
        # Configure TLabel with better font
        style.configure('TLabel', 
                       background=bg_color, 
                       foreground=fg_color,
                       font=('Segoe UI', 9))
        
        # Configure heading labels
        style.configure('Heading.TLabel',
                       font=('Segoe UI', 11, 'bold'),
                       foreground=accent_color)
        
        # Configure TButton with modern styling
        style.configure('TButton',
                       font=('Segoe UI', 9),
                       padding=6)
        
        # Configure primary button
        style.configure('Primary.TButton',
                       font=('Segoe UI', 9, 'bold'))
        
        # Configure TEntry
        style.configure('TEntry',
                       font=('Segoe UI', 9))
        
        # Configure TCombobox
        style.configure('TCombobox',
                       font=('Segoe UI', 9))
        
        # Configure notebook tabs
        style.configure('TNotebook.Tab',
                       font=('Segoe UI', 10, 'bold'),
                       padding=[20, 10])
        
        # Configure LabelFrame
        style.configure('TLabelframe',
                       background=bg_color,
                       font=('Segoe UI', 10, 'bold'))
        style.configure('TLabelframe.Label',
                       background=bg_color,
                       foreground=accent_color,
                       font=('Segoe UI', 10, 'bold'))
        
    def setup_ui(self):
        """Setup the user interface"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(7, weight=1)  # Make row 7 (notebook) expandable
        
        # Project selection section
        ttk.Label(main_frame, text="📁 Project:", style='Heading.TLabel').grid(row=0, column=0, sticky=tk.W, pady=(0, 8), padx=(0, 10))
        project_frame = ttk.Frame(main_frame)
        project_frame.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=(0, 5))
        project_frame.columnconfigure(0, weight=1)
        
        self.project_var = tk.StringVar()
        self.project_combo = ttk.Combobox(project_frame, textvariable=self.project_var, width=50)
        self.project_combo.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 5))
        self.project_combo.bind('<<ComboboxSelected>>', self.on_project_selected)
        self.project_combo.bind('<KeyRelease>', self.filter_projects_on_type)
        self.project_combo.bind('<FocusIn>', self.on_project_focus_in)
        # Remove the Button-1 binding that was interfering with typing
        
        ttk.Button(project_frame, text="Load Projects", command=self.load_projects).grid(row=0, column=1)
        
        # MR filters section
        ttk.Label(main_frame, text="🔍 MR Filter:", style='Heading.TLabel').grid(row=1, column=0, sticky=tk.W, pady=(0, 8), padx=(0, 10))
        filter_frame = ttk.Frame(main_frame)
        filter_frame.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=(0, 5))
        
        self.mr_state_var = tk.StringVar(value="merged")
        ttk.Label(filter_frame, text="State:").pack(side=tk.LEFT, padx=(0, 5))
        state_combo = ttk.Combobox(filter_frame, textvariable=self.mr_state_var, width=15, state="readonly")
        state_combo['values'] = ("all", "opened", "closed", "merged")
        state_combo.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(filter_frame, text="Load MRs", command=self.load_merge_requests).pack(side=tk.LEFT, padx=(0, 5))
        
        # MR selection section
        ttk.Label(main_frame, text="📝 Select MR:", style='Heading.TLabel').grid(row=2, column=0, sticky=tk.W, pady=(0, 8), padx=(0, 10))
        mr_frame = ttk.Frame(main_frame)
        mr_frame.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=(0, 5))
        mr_frame.columnconfigure(0, weight=1)
        
        self.mr_var = tk.StringVar()
        self.mr_combo = ttk.Combobox(mr_frame, textvariable=self.mr_var, width=50)
        self.mr_combo.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 5))
        self.mr_combo.bind('<KeyRelease>', self.filter_mrs_on_type)
        self.mr_combo.bind('<FocusIn>', self.on_mr_focus_in)
        self.mr_combo.bind('<<ComboboxSelected>>', self.on_mr_selected)
        # Remove the Button-1 binding that was interfering with typing
        
        # Alternative: MR URL section
        ttk.Label(main_frame, text="🔗 Or Enter MR URL:", style='Heading.TLabel').grid(row=3, column=0, sticky=tk.W, pady=(0, 8), padx=(0, 10))
        url_frame = ttk.Frame(main_frame)
        url_frame.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=(0, 8))
        url_frame.columnconfigure(0, weight=1)
        
        self.url_entry = ttk.Entry(url_frame, textvariable=self.url_var, width=50, font=('Segoe UI', 9))
        self.url_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 5))
        
        # MR Information section
        info_frame = ttk.LabelFrame(main_frame, text="📊 MR Information", padding=(10, 5))
        info_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(5, 8))
        info_frame.columnconfigure(1, weight=1)
        
        ttk.Label(info_frame, text="📅 Created:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        ttk.Label(info_frame, textvariable=self.mr_created_var, font=('Consolas', 9)).grid(row=0, column=1, sticky=tk.W)
        
        ttk.Label(info_frame, text="✅ Merged:").grid(row=1, column=0, sticky=tk.W, padx=(0, 10), pady=(2, 0))
        ttk.Label(info_frame, textvariable=self.mr_merged_var, font=('Consolas', 9)).grid(row=1, column=1, sticky=tk.W, pady=(2, 0))
        
        ttk.Label(info_frame, text="👥 Assignee Changes:").grid(row=2, column=0, sticky=tk.W, padx=(0, 10), pady=(2, 0))
        self.mr_assignees_label = ttk.Label(info_frame, textvariable=self.mr_assignees_var, font=('Consolas', 9), wraplength=800)
        self.mr_assignees_label.grid(row=2, column=1, sticky=tk.W, pady=(2, 0))
        
        # Buttons section
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=5, column=0, columnspan=2, pady=15)
        
        self.fetch_button = ttk.Button(button_frame, text="🔄 Fetch Comments", command=self.fetch_comments, style='Primary.TButton')
        self.fetch_button.grid(row=0, column=0, padx=(0, 10))
        
        ttk.Button(button_frame, text="🗑️ Clear", command=self.clear_results).grid(row=0, column=1)
        
        # Progress bar
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(5, 5))
        
        # Results section with notebook (tabs)
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=7, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Tab 1: Comments Review
        self.comments_review_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.comments_review_frame, text="Comments Review")
        self.setup_comments_review_tab()
        
        # Tab 2: Best Practices
        self.best_practices_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.best_practices_frame, text="Best Practices")
        self.setup_best_practices_tab()
        
        # Tab 3: Settings
        self.settings_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.settings_frame, text="Settings")
        self.setup_settings_tab()
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("Ready")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.grid(row=8, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
    def setup_comments_review_tab(self):
        """Setup the comments review tab with checkboxes for each discussion"""
        self.comments_review_frame.columnconfigure(0, weight=1)
        self.comments_review_frame.rowconfigure(1, weight=1)
        
        # Control frame for buttons
        control_frame = ttk.Frame(self.comments_review_frame)
        control_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=10, pady=10)
        
        ttk.Button(control_frame, text="✓ Check All", command=self.check_all_comments).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(control_frame, text="✗ Uncheck All", command=self.uncheck_all_comments).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(control_frame, text="🤖 Extract Best Practices", command=self.extract_best_practices, style='Primary.TButton').pack(side=tk.LEFT, padx=(0, 8))
        
        # Scrollable frame for comment blocks
        self.review_canvas = tk.Canvas(self.comments_review_frame)
        self.review_scrollbar = ttk.Scrollbar(self.comments_review_frame, orient="vertical", command=self.review_canvas.yview)
        self.review_scrollable_frame = ttk.Frame(self.review_canvas)
        
        self.review_scrollable_frame.bind(
            "<Configure>",
            lambda e: self.review_canvas.configure(scrollregion=self.review_canvas.bbox("all"))
        )
        
        self.review_canvas.create_window((0, 0), window=self.review_scrollable_frame, anchor="nw")
        self.review_canvas.configure(yscrollcommand=self.review_scrollbar.set)
        
        self.review_canvas.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(5, 0), pady=5)
        self.review_scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S), pady=5)
        
        # Bind mousewheel to canvas
        def _on_mousewheel(event):
            self.review_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        self.review_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Store checkbox variables
        self.comment_checkboxes = {}
        
    def test_token(self):
        """Test the GitLab access token"""
        if not self.gitlab_token:
            messagebox.showerror("Error", "No GitLab token found. Please add it to token.json")
            return
            
        def test_in_thread():
            self.progress.start()
            self.status_var.set("Testing token...")
            
            try:
                api = GitLabAPI(self.gitlab_token)
                success, message = api.test_connection()
                
                if success:
                    messagebox.showinfo("Success", message)
                    self.status_var.set("Token validated successfully")
                else:
                    messagebox.showerror("Error", message)
                    self.status_var.set("Token validation failed")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to test token: {str(e)}")
                self.status_var.set("Token test failed")
            finally:
                self.progress.stop()
        
        threading.Thread(target=test_in_thread, daemon=True).start()
    
    def reset_tabs(self):
        """Reset both Comments Review and Best Practices tabs to initial state"""
        # Clear Comments Review tab
        for widget in self.review_scrollable_frame.winfo_children():
            widget.destroy()
        self.comment_checkboxes.clear()
        
        # Clear Best Practices tab
        self.best_practices_text.config(state="normal")
        self.best_practices_text.delete(1.0, tk.END)
        self.best_practices_text.insert(1.0, "Extract Best Practices from Review Comments\n\n"
                                              "1. Fetch comments from a merge request\n"
                                              "2. Click 'Extract Best Practices' button to analyze the comments\n"
                                              "3. The coding standards will appear here")
        self.best_practices_text.config(state="disabled")
        
        # Clear MR information
        self.mr_created_var.set("")
        self.mr_merged_var.set("")
        self.mr_assignees_var.set("")
        
        # Reset data
        self.comments_data = []
        
    def display_comments(self, discussions):
        """Display comments in the UI
        
        Args:
            discussions: List of discussion objects from GitLab API
            
        Returns:
            int: Number of user discussions displayed
        """
        # Clear comments review tab
        for widget in self.review_scrollable_frame.winfo_children():
            widget.destroy()
        self.comment_checkboxes.clear()
        
        # Populate comments review tab
        return self.populate_comments_review(discussions)
        
    def populate_comments_review(self, discussions):
        """Populate the comments review tab with checkboxes for each discussion
        
        Args:
            discussions: List of discussion objects from GitLab API
            
        Returns:
            int: Number of user discussions displayed
        """
        discussion_count = 0
        skipped_count = 0
        
        for i, discussion in enumerate(discussions):
            notes = discussion.get('notes', [])
            
            # Skip empty discussions
            if not notes:
                skipped_count += 1
                continue
            
            # Filter out system notes (like "marked as resolved", "assigned to", etc.)
            user_notes = [note for note in notes if not note.get('system', False)]
            
            # Skip discussions with only system notes
            if not user_notes:
                skipped_count += 1
                continue
                
            discussion_count += 1
            
            # Create frame for this discussion block
            discussion_frame = ttk.LabelFrame(self.review_scrollable_frame, 
                                            text=f"Discussion {discussion_count}", 
                                            padding="10")
            discussion_frame.pack(fill="x", padx=5, pady=5)
            
            # Create checkbox variable and checkbox (checked by default)
            var = tk.BooleanVar(value=True)
            discussion_id = discussion.get('id', f'discussion_{i}')
            self.comment_checkboxes[discussion_id] = var
            
            checkbox_frame = ttk.Frame(discussion_frame)
            checkbox_frame.pack(fill="x", pady=(0, 10))
            
            checkbox = ttk.Checkbutton(checkbox_frame, 
                                     text=f"Export Discussion {discussion_count} to Comments Repo", 
                                     variable=var)
            checkbox.pack(side=tk.LEFT)
            
            # Add discussion info
            info_frame = ttk.Frame(discussion_frame)
            info_frame.pack(fill="x", pady=(0, 10))
            
            # Check if this is a code comment and get code context
            is_code_comment = False
            position_info = ""
            code_context = None
            
            file_info = get_code_context_from_discussion(discussion)
            if file_info and file_info.get('file_path') != 'Unknown file':
                is_code_comment = True
                position_info = f"📁 File: {file_info['file_path']}"
                if file_info.get('line_number'):
                    position_info += f" (Line {file_info['line_number']})"
                    
                    # Fetch code context from GitLab
                    if hasattr(self, 'current_api') and hasattr(self, 'current_project_id'):
                        try:
                            success, lines_data = self.current_api.get_file_lines_around(
                                self.current_project_id, 
                                file_info['file_path'], 
                                file_info['line_number'],
                                context_lines=3
                            )
                            if success:
                                code_context = lines_data
                        except Exception as e:
                            print(f"Error fetching code context: {e}")
            
            # Add info labels
            if is_code_comment:
                ttk.Label(info_frame, text="💻 Code Comment", foreground="blue").pack(side=tk.LEFT, padx=(0, 10))
            else:
                ttk.Label(info_frame, text="💬 General Comment", foreground="green").pack(side=tk.LEFT, padx=(0, 10))
            
            if position_info:
                ttk.Label(info_frame, text=position_info, foreground="gray").pack(side=tk.LEFT)
            
            # Add code context if available
            if code_context:
                code_frame = ttk.LabelFrame(discussion_frame, text="📄 Code Context", padding="5")
                code_frame.pack(fill="x", pady=(5, 10))
                
                # Create text widget for code display
                code_text = tk.Text(code_frame, wrap=tk.NONE, height=len(code_context['lines']) + 1, 
                                  width=100, font=("Consolas", 9), relief="sunken", borderwidth=1)
                code_text.pack(fill="x", padx=5, pady=5)
                
                # Add horizontal scrollbar for long lines
                code_scrollbar = ttk.Scrollbar(code_frame, orient="horizontal", command=code_text.xview)
                code_text.configure(xscrollcommand=code_scrollbar.set)
                code_scrollbar.pack(fill="x", padx=5)
                
                # Configure text tags for highlighting
                code_text.tag_configure("target_line", background="#ffeb3b", foreground="#000")
                code_text.tag_configure("line_number", foreground="#666", font=("Consolas", 8))
                code_text.tag_configure("code_content", font=("Consolas", 9))
                
                # Insert code lines
                for line_data in code_context['lines']:
                    line_num = line_data['number']
                    content = line_data['content']
                    is_target = line_data['is_target']
                    
                    # Format line number (right-aligned in 4 characters)
                    line_num_str = f"{line_num:4d}: "
                    code_text.insert(tk.END, line_num_str, "line_number")
                    
                    # Insert code content
                    if is_target:
                        code_text.insert(tk.END, content + "\n", ("code_content", "target_line"))
                    else:
                        code_text.insert(tk.END, content + "\n", "code_content")
                
                code_text.config(state="disabled")  # Make read-only
            
            # Add comments in this discussion
            for note_idx, note in enumerate(user_notes):
                author = note.get('author', {}).get('name', 'Unknown')
                created_at = format_datetime(note.get('created_at', ''))
                body = extract_comment_text(note)
                
                # Create frame for this comment
                comment_frame = ttk.Frame(discussion_frame)
                comment_frame.pack(fill="x", pady=5)
                
                # Author and date header
                header = f"👤 {author} • 📅 {created_at}"
                header_label = ttk.Label(comment_frame, text=header, font=("TkDefaultFont", 9, "bold"))
                header_label.pack(anchor="w")
                
                # Comment body
                comment_text = tk.Text(comment_frame, wrap=tk.WORD, height=4, width=80, 
                                     relief="groove", borderwidth=1, padx=5, pady=5)
                comment_text.pack(fill="x", pady=(5, 0))
                comment_text.insert("1.0", body)
                comment_text.config(state="disabled")  # Make read-only
                
                # Add separator between comments (except for last comment)
                if note_idx < len(user_notes) - 1:
                    separator = ttk.Separator(discussion_frame, orient='horizontal')
                    separator.pack(fill="x", pady=(5, 0))
        
        # Show summary message if discussions were skipped
        if skipped_count > 0:
            summary_frame = ttk.Frame(self.review_scrollable_frame)
            summary_frame.pack(fill="x", padx=5, pady=10)
            
            summary_text = f"ℹ️ Showing {discussion_count} user discussions. "
            summary_text += f"{skipped_count} system-generated discussion(s) were filtered out."
            
            ttk.Label(summary_frame, text=summary_text, foreground="gray", 
                     font=("TkDefaultFont", 9, "italic")).pack(anchor="w")
        
        # If no discussions at all, show a message
        if discussion_count == 0:
            no_comments_frame = ttk.Frame(self.review_scrollable_frame)
            no_comments_frame.pack(fill="both", expand=True, padx=20, pady=20)
            
            ttk.Label(no_comments_frame, 
                     text="No user comments found in this merge request.",
                     font=("TkDefaultFont", 11)).pack(pady=10)
            ttk.Label(no_comments_frame, 
                     text=f"Total discussions fetched: {len(discussions)}",
                     foreground="gray").pack()
            ttk.Label(no_comments_frame, 
                     text=f"System-generated discussions filtered: {skipped_count}",
                     foreground="gray").pack()
            ttk.Label(no_comments_frame, 
                     text="\nThis MR may only contain automated system messages\n(e.g., 'marked as resolved', 'assigned to', merge status updates).",
                     foreground="gray", justify="center").pack(pady=10)
        
        return discussion_count
    
    def export_comments(self):
        """Export comments to JSON file"""
        if not self.comments_data:
            messagebox.showerror("Error", "No comments data to export")
            return
            
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Save comments as JSON"
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(self.comments_data, f, indent=2, ensure_ascii=False)
                messagebox.showinfo("Success", f"Comments exported to {filename}")
                self.status_var.set(f"Exported to {filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export: {str(e)}")
                self.status_var.set("Export failed")
    
    def on_standards_type_changed(self, event=None):
        """Handle standards type dropdown selection change"""
        selected = self.standards_type_var.get()
        
        if selected == "Custom Link":
            # Enable custom link entry for user input
            self.excel_link_entry.config(state="normal")
            self.excel_link_var.set("")
        else:
            # Set the predefined link and disable entry
            self.excel_link_entry.config(state="disabled")
            self.excel_link_var.set(self.standards_links[selected])
    
    def copy_standards_to_clipboard(self):
        """Copy coding standards to clipboard for pasting into SharePoint/Teams Excel"""
        content = self.best_practices_text.get(1.0, tk.END).strip()
        
        if not content or "Extract Best Practices from Review Comments" in content:
            messagebox.showwarning("Warning", "No coding standards to copy. Please extract best practices first.")
            return
        
        try:
            # Extract only the actual standards content
            # Skip header lines and remove sequence numbers
            lines = content.split('\n')
            cleaned_lines = []
            skip_header = True
            
            for line in lines:
                stripped = line.strip()
                
                # Skip header section (until we find the actual content)
                if skip_header:
                    if stripped.startswith('Here are the coding standards') or \
                       stripped.startswith('Based on the review') or \
                       '=' in stripped or \
                       'Extracted Coding Standards' in stripped or \
                       'Generated by' in stripped or \
                       not stripped:
                        continue
                    else:
                        skip_header = False
                
                # Remove sequence numbers (e.g., "1.", "2.", etc.) at the start of lines
                import re
                cleaned = re.sub(r'^\d+\.\s*', '', stripped)
                
                # Keep all lines including empty ones (for spacing between standards)
                cleaned_lines.append(cleaned)
            
            cleaned_content = '\n'.join(cleaned_lines)
            
            if not cleaned_content:
                messagebox.showwarning("Warning", "No coding standards found to copy.")
                return
            
            # Copy to clipboard
            self.root.clipboard_clear()
            self.root.clipboard_append(cleaned_content)
            self.root.update()  # Required for clipboard to work
            
            excel_link = self.excel_link_var.get().strip()
            
            if excel_link:
                msg = (
                    "✅ Coding standards copied to clipboard!\n\n"
                    "Next steps:\n"
                    "1. Click the link below to open your SharePoint/Teams Excel\n"
                    "2. Select cell A1 (or desired starting cell)\n"
                    "3. Press Ctrl+V to paste\n\n"
                    f"Link: {excel_link[:80]}..."
                )
                
                # Ask if user wants to open the link
                result = messagebox.askokcancel("Copied to Clipboard", msg + "\n\nOpen link in browser?")
                if result:
                    import webbrowser
                    webbrowser.open(excel_link)
            else:
                messagebox.showinfo("Copied to Clipboard", 
                    "✅ Coding standards copied to clipboard!\n\n"
                    "You can now paste them into your Excel file.")
            
            self.status_var.set("Coding standards copied to clipboard")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to copy to clipboard: {str(e)}")
            self.status_var.set("Copy failed")
    
    def export_to_excel(self):
        """Export best practices to local Excel file"""
        # Get the content from the text area
        content = self.best_practices_text.get(1.0, tk.END).strip()
        
        if not content or "Extract Best Practices from Review Comments" in content:
            messagebox.showwarning("Warning", "No coding standards to export. Please extract best practices first.")
            return
        
        # Get the Excel link (for reference in the file)
        excel_link = self.excel_link_var.get().strip()
        
        # Export to local file
        self._export_to_local_excel(content, excel_link)
    
    def _export_to_local_excel(self, content, excel_link):
        """Export to local Excel file"""
        # Ask where to save the file
        default_name = "Coding_Standards.xlsx"
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save coding standards as Excel",
            initialfile=default_name
        )
        
        if not filename:
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Coding Standards"
            
            # Set column widths
            ws.column_dimensions['A'].width = 100
            
            # Add title
            ws['A1'] = "Extracted Coding Standards"
            ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30
            
            # Add SharePoint/Teams link if provided
            current_row = 3
            if excel_link:
                ws[f'A{current_row}'] = f"📁 SharePoint/Teams Location:"
                ws[f'A{current_row}'].font = Font(bold=True, size=10)
                current_row += 1
                
                ws[f'A{current_row}'] = excel_link
                ws[f'A{current_row}'].font = Font(size=9, color="0563C1", underline="single")
                ws[f'A{current_row}'].alignment = Alignment(wrap_text=True)
                current_row += 1
                
                ws[f'A{current_row}'] = "ℹ️  Upload this file to the above location"
                ws[f'A{current_row}'].font = Font(italic=True, size=9, color="666666")
                current_row += 2
            
            # Add instructions
            ws[f'A{current_row}'] = "📋 How to Upload:"
            ws[f'A{current_row}'].font = Font(bold=True, size=10)
            current_row += 1
            
            ws[f'A{current_row}'] = "1. Open the SharePoint/Teams link above in your browser"
            ws[f'A{current_row}'].font = Font(size=9)
            current_row += 1
            
            ws[f'A{current_row}'] = "2. Click 'Upload' and select this file, OR copy the standards below and paste into the existing file"
            ws[f'A{current_row}'].font = Font(size=9)
            ws[f'A{current_row}'].alignment = Alignment(wrap_text=True)
            current_row += 2
            
            # Add separator
            ws[f'A{current_row}'] = "─" * 120
            ws[f'A{current_row}'].font = Font(bold=True, color="4472C4")
            current_row += 2
            
            # Parse and add content
            lines = content.split('\n')
            
            for line in lines:
                if line.strip():
                    ws[f'A{current_row}'] = line
                    ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Style headers (lines with = or -)
                    if line.strip().startswith('='):
                        ws[f'A{current_row}'].font = Font(bold=True, size=12, color="4472C4")
                        ws.row_dimensions[current_row].height = 20
                    elif line.strip().startswith('-'):
                        ws[f'A{current_row}'].font = Font(bold=True, size=10)
                    # Style bullet points
                    elif line.strip().startswith('•'):
                        ws[f'A{current_row}'].font = Font(size=10)
                        ws[f'A{current_row}'].alignment = Alignment(wrap_text=True, indent=1)
                    else:
                        ws[f'A{current_row}'].font = Font(size=10)
                    
                    current_row += 1
            
            # Save the workbook
            wb.save(filename)
            
            if excel_link:
                msg = (
                    f"✅ Coding standards exported to:\n{filename}\n\n"
                    f"📁 Next steps:\n"
                    f"1. Open the SharePoint/Teams link in your browser\n"
                    f"2. Upload this Excel file, OR\n"
                    f"3. Copy the standards from the file and paste into the online Excel\n\n"
                    f"The SharePoint/Teams link is included in the Excel file for reference."
                )
            else:
                msg = f"✅ Coding standards exported to:\n{filename}\n\nYou can now upload this file to SharePoint/Teams."
            
            messagebox.showinfo("Export Successful", msg)
            self.status_var.set(f"Exported to {filename}")
            
        except ImportError:
            messagebox.showerror("Error", 
                "openpyxl library is required for Excel export.\n"
                "Please install it using: pip install openpyxl")
            self.status_var.set("Export failed - missing openpyxl")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {str(e)}")
            self.status_var.set("Export failed")
    
    def view_images(self):
        """Open image viewer window to display downloaded images"""
        if not self.downloaded_images:
            messagebox.showinfo("No Images", "No images were found in the comments.")
            return
        
        image_paths = list(self.downloaded_images.values())
        self.image_viewer.create_image_display_window(
            image_paths, 
            f"Images from Merge Request Comments ({len(image_paths)} images)"
        )
    
    def check_all_comments(self):
        """Check all comment checkboxes"""
        for var in self.comment_checkboxes.values():
            var.set(True)
    
    def uncheck_all_comments(self):
        """Uncheck all comment checkboxes"""
        for var in self.comment_checkboxes.values():
            var.set(False)
    
    def export_checked_comments(self):
        """Export only the checked comments to JSON"""
        if not self.comments_data:
            messagebox.showerror("Error", "No comments data to export")
            return
        
        # Get checked discussions
        checked_discussions = []
        for discussion_id, var in self.comment_checkboxes.items():
            if var.get():
                # Find the discussion in comments_data
                for discussion in self.comments_data:
                    if discussion.get('id') == discussion_id:
                        checked_discussions.append(discussion)
                        break
        
        if not checked_discussions:
            messagebox.showwarning("Warning", "No comments are checked for export")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Save checked comments as JSON"
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(checked_discussions, f, indent=2, ensure_ascii=False)
                messagebox.showinfo("Success", f"Checked comments exported to {filename}")
                self.status_var.set(f"Exported {len(checked_discussions)} checked discussions")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export: {str(e)}")
                self.status_var.set("Export failed")
    
    def clear_results(self):
        """Clear all results and reset the interface"""
        self.all_comments_text.delete(1.0, tk.END)
        self.code_comments_text.delete(1.0, tk.END)
        self.summary_text.delete(1.0, tk.END)
        
        # Clear comments review tab
        for widget in self.review_scrollable_frame.winfo_children():
            widget.destroy()
        self.comment_checkboxes.clear()
        
        # Clear MR information
        self.mr_created_var.set("")
        self.mr_merged_var.set("")
        self.mr_assignees_var.set("")
        
        self.comments_data = None
        self.current_api = None
        self.current_project_id = None
        self.status_var.set("Results cleared")
    
    def test_button_click(self):
        """Test method to verify button clicks work"""
        print("DEBUG: test_button_click called!")
        with open("debug.log", "a") as f:
            f.write("DEBUG: test_button_click called!\n")
        messagebox.showinfo("Test", "Button click works!")
        
    def load_projects(self):
        """Load user's projects from GitLab API"""
        print("DEBUG: load_projects called")
        if not self.gitlab_token:
            print("DEBUG: No token found")
            messagebox.showwarning("Warning", "No GitLab token found. Please add it to token.json")
            return
        
        # Reset tabs when loading new projects
        self.reset_tabs()
        
        # Clear MR URL field
        self.url_var.set("")
        
        def load_in_thread():
            self.progress.start()
            self.status_var.set("Loading Certificate-forms platform projects...")
            
            try:
                api = GitLabAPI(self.gitlab_token)
                success, projects = api.get_user_projects()
                
                print(f"DEBUG: API call result - success: {success}")
                if success:
                    print(f"DEBUG: Found {len(projects) if projects else 0} projects")
                    self.status_var.set(f"Processing {len(projects)} projects...")
                    # Convert GitLab projects to our format
                    self.projects_data = []
                    for proj in projects:
                        project_data = {
                            'name': proj.get('name', 'Unknown'),
                            'path': proj.get('path_with_namespace', ''),
                            'description': proj.get('description', ''),
                            'id': proj.get('id'),
                            'web_url': proj.get('web_url', ''),
                            'last_activity_at': proj.get('last_activity_at', ''),
                            'visibility': proj.get('visibility', 'private')
                        }
                        self.projects_data.append(project_data)
                    
                    # Create display names
                    project_names = []
                    for proj in self.projects_data:
                        # Format: "Project Name (path) - visibility"
                        name = f"{proj['name']} ({proj['path']}) - {proj['visibility']}"
                        project_names.append(name)
                    
                    self.all_project_names = project_names.copy()
                    self.project_combo['values'] = project_names
                    # Enable typing in combobox for search
                    self.project_combo['state'] = 'normal'
                    
                    if project_names:
                        self.status_var.set(f"Loaded {len(project_names)} Certificate-forms platform projects")
                    else:
                        self.status_var.set("No Certificate-forms platform projects found")
                else:
                    messagebox.showerror("Error", f"Failed to load Certificate-forms projects: {projects}")
                    self.status_var.set("Failed to load Certificate-forms projects")
                    
            except Exception as e:
                messagebox.showerror("Error", f"Unexpected error: {str(e)}")
                self.status_var.set("Error loading Certificate-forms projects")
            finally:
                self.progress.stop()
        
        print("DEBUG: Starting thread")
        threading.Thread(target=load_in_thread, daemon=True).start()
    
    def filter_projects_on_type(self, event=None):
        """Filter projects as user types in the combobox"""
        if self.is_filtering:
            return
            
        self.is_filtering = True
        try:
            current_text = self.project_var.get().lower()
            
            if not current_text:
                # Show all projects if search is empty
                filtered_projects = self.all_project_names
            else:
                # Filter projects that contain the search text
                filtered_projects = [
                    name for name in self.all_project_names 
                    if current_text in name.lower()
                ]
            
            # Update combobox values without forcing dropdown open
            self.project_combo['values'] = filtered_projects
                
        finally:
            self.is_filtering = False
    
    def on_project_focus_in(self, event=None):
        """Handle project combobox focus to show all projects if none are filtered"""
        if not self.project_combo['values'] and self.all_project_names:
            self.project_combo['values'] = self.all_project_names
    
    def filter_mrs_on_type(self, event=None):
        """Filter MRs as user types in the combobox"""
        if self.is_filtering_mrs:
            return
            
        self.is_filtering_mrs = True
        try:
            current_text = self.mr_var.get().lower()
            
            if not current_text:
                # Show all MRs if search is empty
                filtered_mrs = self.all_mr_names
            else:
                # Filter MRs that contain the search text
                filtered_mrs = [
                    name for name in self.all_mr_names 
                    if current_text in name.lower()
                ]
            
            # Update combobox values without forcing dropdown open
            self.mr_combo['values'] = filtered_mrs
                
        finally:
            self.is_filtering_mrs = False
    
    def on_mr_focus_in(self, event=None):
        """Handle MR combobox focus to show all MRs if none are filtered"""
        if not self.mr_combo['values'] and self.all_mr_names:
            self.mr_combo['values'] = self.all_mr_names
    
    def on_project_selected(self, event=None):
        """Handle project selection"""
        selected = self.project_combo.current()
        if selected >= 0 and selected < len(self.projects_data):
            project = self.projects_data[selected]
            self.status_var.set(f"Selected: {project['name']}")
            # Clear MR selection and search data when project changes
            self.mr_combo.set('')
            self.mr_combo['values'] = []
            self.all_mr_names = []
            self.current_mrs = []
            self.mr_combo['state'] = 'readonly'
    
    def load_merge_requests(self):
        """Load merge requests for the selected project"""
        print("DEBUG: load_merge_requests called")
        selected = self.project_combo.current()
        print(f"DEBUG: Selected project index: {selected}")
        print(f"DEBUG: Projects data length: {len(self.projects_data)}")
        
        if selected < 0 or selected >= len(self.projects_data):
            messagebox.showwarning("Warning", "Please select a project first")
            return
        
        token = self.gitlab_token
        if not token:
            messagebox.showerror("Error", "GitLab token not found. Please add it to token.json")
            return
        
        project = self.projects_data[selected]
        project_path = project['path']
        mr_state = self.mr_state_var.get()
        
        print(f"DEBUG: Loading MRs for project: {project['name']}")
        print(f"DEBUG: Project path: {project_path}")
        print(f"DEBUG: MR state filter: {mr_state}")
        
        def load_in_thread():
            self.progress.start()
            self.status_var.set(f"Loading {mr_state} merge requests from {project['name']}...")
            
            try:
                api = GitLabAPI(token)
                print(f"DEBUG: Calling get_merge_requests API...")
                success, mrs = api.get_merge_requests(project_path, state=mr_state)
                print(f"DEBUG: API response - success: {success}, MR count: {len(mrs) if success else 'N/A'}")
                
                if success:
                    self.current_mrs = mrs
                    mr_options = []
                    
                    for mr in mrs:
                        title = mr.get('title', 'No title')
                        iid = mr.get('iid', 'N/A')
                        state = mr.get('state', 'unknown')
                        author = mr.get('author', {}).get('name', 'Unknown')
                        created_at = mr.get('created_at', '')
                        updated_at = mr.get('updated_at', '')
                        
                        # Format date (prefer created_at for consistency)
                        try:
                            if created_at:
                                date_part = created_at.split('T')[0]
                            elif updated_at:
                                date_part = updated_at.split('T')[0]
                            else:
                                date_part = 'N/A'
                        except:
                            date_part = 'N/A'
                        
                        # Create display text with creation date for better chronological sorting
                        display_text = f"MR!{iid} - {title} ({state}) - {author} - {date_part}"
                        mr_options.append(display_text)
                    
                    self.all_mr_names = mr_options.copy()
                    self.mr_combo['values'] = mr_options
                    # Enable typing in combobox for search
                    self.mr_combo['state'] = 'normal'
                    self.status_var.set(f"Loaded {len(mrs)} {mr_state} merge requests")
                else:
                    messagebox.showerror("Error", f"Failed to load MRs: {mrs}")
                    self.status_var.set("Failed to load merge requests")
            except Exception as e:
                messagebox.showerror("Error", f"Unexpected error: {str(e)}")
                self.status_var.set("Error loading merge requests")
            finally:
                self.progress.stop()
        
        threading.Thread(target=load_in_thread, daemon=True).start()
    
    def on_mr_selected(self, event=None):
        """Handle MR selection"""
        selected_text = self.mr_var.get()
        if not selected_text:
            return
        
        # Extract MR IID from the selected text (format: "MR!{iid} - ...")
        try:
            import re
            match = re.search(r'MR!(\d+)', selected_text)
            if not match:
                return
            
            selected_iid = int(match.group(1))
            
            # Find the MR with matching IID
            mr = None
            for m in self.current_mrs:
                if m.get('iid') == selected_iid:
                    mr = m
                    break
            
            if not mr:
                self.status_var.set(f"Error: Could not find MR!{selected_iid}")
                return
            
            project = self.projects_data[self.project_combo.current()]
            
            # Construct GitLab URL (project['path'] already contains path_with_namespace)
            mr_url = f"https://gitlab.com/{project['path']}/-/merge_requests/{mr['iid']}"
            self.url_var.set(mr_url)
            
            self.status_var.set(f"Selected MR!{mr['iid']}: {mr['title']}")
        except Exception as e:
            self.status_var.set(f"Error selecting MR: {str(e)}")
    
    def update_mr_information(self, project_path, mr_iid):
        """Fetch and display MR creation, merge, and assignee change dates
        
        Args:
            project_path (str): GitLab project path
            mr_iid (int): Merge request internal ID
        """
        try:
            api = GitLabAPI(self.gitlab_token)
            
            # Get MR details
            success, mr_data = api.get_merge_request_details(project_path, mr_iid)
            
            if success:
                # Created date
                created_at = mr_data.get('created_at', '')
                if created_at:
                    self.mr_created_var.set(format_datetime(created_at))
                else:
                    self.mr_created_var.set("N/A")
                
                # Merged date
                merged_at = mr_data.get('merged_at', '')
                if merged_at:
                    self.mr_merged_var.set(format_datetime(merged_at))
                else:
                    state = mr_data.get('state', 'unknown')
                    if state == 'merged':
                        self.mr_merged_var.set("Merged (date unavailable)")
                    else:
                        self.mr_merged_var.set(f"Not merged ({state})")
            else:
                self.mr_created_var.set("Failed to fetch")
                self.mr_merged_var.set("Failed to fetch")
                print(f"Failed to get MR details: {mr_data}")
            
            # Get system notes to find assignee change events
            success_notes, notes = api.get_merge_request_notes(project_path, mr_iid)
            
            assignee_changes = []
            if success_notes:
                # Parse system notes for assignee changes
                for note in notes:
                    if note.get('system', False):
                        body = note.get('body', '')
                        created_at = note.get('created_at', '')
                        author = note.get('author', {}).get('name', 'Unknown')
                        
                        # Look for assignee-related system messages
                        # GitLab uses phrases like "assigned to @username" or "unassigned @username"
                        if 'assigned to @' in body.lower() or 'unassigned @' in body.lower() or 'changed assignee' in body.lower():
                            assignee_changes.append({
                                'date': created_at,
                                'action': body,
                                'by': author
                            })
                
                # Display assignee information
                if success and mr_data:
                    assignees = mr_data.get('assignees', [])
                    assignee_info = []
                    
                    if assignees:
                        for assignee in assignees:
                            name = assignee.get('name', 'Unknown')
                            assignee_info.append(name)
                        
                        # Build assignee string with change history
                        reviewers = mr_data.get('reviewers', [])
                        assignee_str_parts = []
                        
                        if assignee_info:
                            assignee_str_parts.append(f"Current: {', '.join(assignee_info)}")
                        
                        if reviewers:
                            reviewer_names = [r.get('name', 'Unknown') for r in reviewers]
                            assignee_str_parts.append(f"Reviewers: {', '.join(reviewer_names)}")
                        
                        # Add last assignment change date
                        if assignee_changes:
                            last_change = assignee_changes[-1]
                            change_date = format_datetime(last_change['date'])
                            assignee_str_parts.append(f"Last changed: {change_date}")
                        
                        self.mr_assignees_var.set(" | ".join(assignee_str_parts))
                    else:
                        if assignee_changes:
                            last_change = assignee_changes[-1]
                            change_date = format_datetime(last_change['date'])
                            self.mr_assignees_var.set(f"No current assignees (Last change: {change_date})")
                        else:
                            self.mr_assignees_var.set("No assignees")
                else:
                    self.mr_assignees_var.set("No assignees")
            else:
                self.mr_assignees_var.set("Failed to fetch assignee data")
                print(f"Failed to get notes: {notes}")
                
        except Exception as e:
            self.mr_created_var.set("Error")
            self.mr_merged_var.set("Error")
            self.mr_assignees_var.set("Error fetching data")
            print(f"Error updating MR information: {str(e)}")
    
    def fetch_comments(self):
        """Fetch comments from the merge request"""
        token = self.gitlab_token
        url = self.url_var.get().strip()
        
        if not token:
            messagebox.showerror("Error", "GitLab token not found. Please add it to token.json")
            return
        
        if not url:
            # Check if we have a selected MR
            if self.mr_combo.current() >= 0:
                self.on_mr_selected()  # Update URL from selected MR
                url = self.url_var.get().strip()
            
            if not url:
                messagebox.showerror("Error", "Please select an MR from dropdown or enter MR URL")
                return
        
        # Parse URL
        success, project_id, mr_iid, error = parse_gitlab_url(url)
        if not success:
            messagebox.showerror("Error", f"Invalid URL: {error}")
            return
            
        def fetch_in_thread():
            # Reset tabs at the start of fetching
            self.reset_tabs()
            
            self.progress.start()
            self.fetch_button.config(state="disabled")
            self.status_var.set("Fetching comments and MR details...")
            
            try:
                api = GitLabAPI(token)
                
                # Fetch MR information (dates, assignees)
                self.update_mr_information(project_id, mr_iid)
                
                success, data, numeric_project_id = api.get_merge_request_discussions(project_id, mr_iid)
                
                if success:
                    # Store references for code context fetching
                    self.current_api = api
                    self.current_project_id = project_id
                    
                    self.comments_data = data
                    
                    # Display comments
                    user_discussion_count = self.display_comments(data)
                    
                    # Status message based on user discussions found
                    if user_discussion_count == 0:
                        self.status_var.set(f"No user comments found ({len(data)} system discussions filtered)")
                    else:
                        self.status_var.set(f"Fetched {user_discussion_count} user discussions successfully")
                else:
                    messagebox.showerror("Error", f"Failed to fetch comments: {data}")
                    self.status_var.set("Failed to fetch comments")
            except Exception as e:
                messagebox.showerror("Error", f"Unexpected error: {str(e)}")
                self.status_var.set("Error occurred")
            finally:
                self.progress.stop()
                self.fetch_button.config(state="normal")
        
        threading.Thread(target=fetch_in_thread, daemon=True).start()
    
    def setup_best_practices_tab(self):
        """Setup the best practices tab"""
        # Create top frame for export controls
        export_frame = ttk.Frame(self.best_practices_frame)
        export_frame.pack(fill=tk.X, padx=10, pady=(10, 5))
        
        # Dropdown for selecting standards type
        ttk.Label(export_frame, text="Standards Type:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.standards_type_var = tk.StringVar()
        self.standards_links = {
            "UI Standards (React)": "https://vertafore.sharepoint.com/:x:/r/teams/HYD-AgencyTeams-Firefoxes/Shared%20Documents/Coding%20Standards/Checklist_CodeReview_React_Dev.xlsx?d=w97f829117543423292258d294de21e4d&csf=1&web=1&e=aWV004",
            "Java Standards": "https://vertafore.sharepoint.com/:x:/r/teams/HYD-AgencyTeams-Firefoxes/Shared%20Documents/Coding%20Standards/Checklist_codeReview-Java-dev.xlsx?d=w1deabc2283894385abf2d0a866301e30&csf=1&web=1&e=QVyyaz",
            "Custom Link": ""
        }
        
        standards_combo = ttk.Combobox(
            export_frame, 
            textvariable=self.standards_type_var, 
            values=list(self.standards_links.keys()),
            state="readonly",
            width=25
        )
        standards_combo.pack(side=tk.LEFT, padx=(0, 10))
        standards_combo.current(0)  # Set default to UI Standards
        standards_combo.bind('<<ComboboxSelected>>', self.on_standards_type_changed)
        
        # Entry for custom link (initially hidden/disabled)
        self.excel_link_var = tk.StringVar()
        self.excel_link_entry = ttk.Entry(export_frame, textvariable=self.excel_link_var, width=35, state="disabled")
        self.excel_link_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        # Set initial link
        self.excel_link_var.set(self.standards_links["UI Standards (React)"])
        
        ttk.Button(export_frame, text="📋 Copy Standards", command=self.copy_standards_to_clipboard).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(export_frame, text="📊 Export to Excel", command=self.export_to_excel, style='Primary.TButton').pack(side=tk.LEFT)
        
        # Main scrollable text area (editable)
        self.best_practices_text = scrolledtext.ScrolledText(
            self.best_practices_frame, 
            wrap=tk.WORD, 
            width=90, 
            height=30,
            font=('Segoe UI', 10),
            padx=10,
            pady=10
        )
        self.best_practices_text.pack(fill=tk.BOTH, expand=True, padx=15, pady=(10, 15))
        
        # Initially show instructions
        instructions = """Extract Best Practices from Review Comments

Instructions:
1. Go to the 'Comments Review' tab
2. Check the discussions you want to analyze
3. Click 'Extract Best Practices' button
4. Make sure you have set your Vertafore API key in llm_token.json

The AI (Claude Sonnet 3.5 via Vertafore API) will analyze the selected review comments and extract:
• Code quality standards
• Best practices for coding
• Security considerations
• Performance recommendations
• Maintainability guidelines
• Testing practices
• Documentation standards

Using Vertafore's enterprise AI platform for secure, compliant analysis.

Note: The extracted standards are editable - you can add, modify, or delete any content.
"""
        self.best_practices_text.insert(tk.END, instructions)
    
    def setup_settings_tab(self):
        """Setup the settings tab for token management"""
        # Main container with canvas for better scrolling
        canvas = tk.Canvas(self.settings_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.settings_frame, orient="vertical", command=canvas.yview)
        settings_container = ttk.Frame(canvas)
        
        settings_container.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=settings_container, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=20, pady=20)
        scrollbar.pack(side="right", fill="y")
        
        # Title with icon
        title_frame = ttk.Frame(settings_container)
        title_frame.pack(pady=(0, 30))
        
        title_label = ttk.Label(title_frame, text="⚙️ API Token Configuration", 
                               font=('Segoe UI', 16, 'bold'))
        title_label.pack()
        
        subtitle_label = ttk.Label(title_frame, text="Configure your access tokens for GitLab and Vertafore AI", 
                                  font=('Segoe UI', 9), foreground="gray")
        subtitle_label.pack(pady=(5, 0))
        
        # GitLab Token Section
        gitlab_frame = ttk.LabelFrame(settings_container, text="  🔑 GitLab Personal Access Token  ", padding="20")
        gitlab_frame.pack(fill=tk.X, pady=(0, 20))
        
        gitlab_desc = ttk.Label(gitlab_frame, 
                               text="Required to authenticate with GitLab API and fetch merge request comments",
                               font=('Segoe UI', 9), foreground="#666666", wraplength=600)
        gitlab_desc.grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 15))
        
        ttk.Label(gitlab_frame, text="Access Token:", font=('Segoe UI', 9, 'bold')).grid(row=1, column=0, sticky=tk.W, pady=8)
        self.gitlab_token_var = tk.StringVar(value=self.gitlab_token if self.gitlab_token else "")
        gitlab_token_entry = ttk.Entry(gitlab_frame, textvariable=self.gitlab_token_var, width=70, show="*", font=('Consolas', 9))
        gitlab_token_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(15, 0), pady=8)
        
        # Show/Hide token button
        self.gitlab_show_var = tk.BooleanVar(value=False)
        show_gitlab_btn = ttk.Checkbutton(gitlab_frame, text="Show token", variable=self.gitlab_show_var,
                                         command=lambda: gitlab_token_entry.config(show="" if self.gitlab_show_var.get() else "*"))
        show_gitlab_btn.grid(row=2, column=1, sticky=tk.W, padx=(15, 0))
        
        gitlab_button_frame = ttk.Frame(gitlab_frame)
        gitlab_button_frame.grid(row=3, column=1, sticky=tk.W, padx=(15, 0), pady=(15, 5))
        
        save_gitlab_btn = ttk.Button(gitlab_button_frame, text="💾 Save Token", 
                                     command=self.save_gitlab_token, width=20)
        save_gitlab_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        test_gitlab_btn = ttk.Button(gitlab_button_frame, text="🔌 Test Connection", 
                                    command=self.test_token, width=20)
        test_gitlab_btn.pack(side=tk.LEFT)
        
        gitlab_frame.columnconfigure(1, weight=1)
        
        storage_label = ttk.Label(gitlab_frame, text="📁 Stored in: token.json", 
                                 foreground="#888888", font=('Segoe UI', 8, 'italic'))
        storage_label.grid(row=4, column=1, sticky=tk.W, padx=(15, 0), pady=(5, 0))
        
        # LLM Token Section
        llm_frame = ttk.LabelFrame(settings_container, text="  🤖 Vertafore Enterprise AI Token  ", padding="20")
        llm_frame.pack(fill=tk.X, pady=(0, 20))
        
        llm_desc = ttk.Label(llm_frame, 
                            text="Required to access Claude Sonnet 3.5 via Vertafore API for extracting coding standards",
                            font=('Segoe UI', 9), foreground="#666666", wraplength=600)
        llm_desc.grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 15))
        
        ttk.Label(llm_frame, text="API Token:", font=('Segoe UI', 9, 'bold')).grid(row=1, column=0, sticky=tk.W, pady=8)
        self.llm_token_var = tk.StringVar(value=self.llm_token if self.llm_token else "")
        llm_token_entry = ttk.Entry(llm_frame, textvariable=self.llm_token_var, width=70, show="*", font=('Consolas', 9))
        llm_token_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(15, 0), pady=8)
        
        # Show/Hide token button
        self.llm_show_var = tk.BooleanVar(value=False)
        show_llm_btn = ttk.Checkbutton(llm_frame, text="Show token", variable=self.llm_show_var,
                                       command=lambda: llm_token_entry.config(show="" if self.llm_show_var.get() else "*"))
        show_llm_btn.grid(row=2, column=1, sticky=tk.W, padx=(15, 0))
        
        save_llm_btn = ttk.Button(llm_frame, text="💾 Save Token", 
                                 command=self.save_llm_token, width=20)
        save_llm_btn.grid(row=3, column=1, sticky=tk.W, padx=(15, 0), pady=(15, 5))
        
        llm_frame.columnconfigure(1, weight=1)
        
        storage_label2 = ttk.Label(llm_frame, text="📁 Stored in: llm_token.json", 
                                  foreground="#888888", font=('Segoe UI', 8, 'italic'))
        storage_label2.grid(row=4, column=1, sticky=tk.W, padx=(15, 0), pady=(5, 0))
        
        # Info Section with better styling
        info_frame = ttk.LabelFrame(settings_container, text="  ℹ️ Information  ", padding="20")
        info_frame.pack(fill=tk.X, pady=(0, 20))
        
        info_items = [
            ("🔐 Security:", "Both tokens are stored locally in encrypted JSON files and never sent to external services"),
            ("🔄 Auto-load:", "Tokens are loaded automatically when the application starts"),
            ("🌐 GitLab Token:", "Used for accessing GitLab API to fetch merge request discussions and comments"),
            ("🧠 AI Token:", "Used for Claude Sonnet 3.5 API calls to analyze and extract coding best practices"),
        ]
        
        for i, (title, desc) in enumerate(info_items):
            item_frame = ttk.Frame(info_frame)
            item_frame.pack(fill=tk.X, pady=(0, 10))
            
            ttk.Label(item_frame, text=title, font=('Segoe UI', 9, 'bold'), 
                     foreground="#0066cc").pack(anchor=tk.W)
            ttk.Label(item_frame, text=desc, font=('Segoe UI', 9), 
                     foreground="#555555", wraplength=650).pack(anchor=tk.W, padx=(20, 0), pady=(2, 0))
    
    def save_gitlab_token(self):
        """Save GitLab token to file"""
        token = self.gitlab_token_var.get().strip()
        if not token:
            messagebox.showwarning("Warning", "Please enter a GitLab token")
            return
        
        try:
            self.token_manager.save_token(token)
            self.gitlab_token = token
            messagebox.showinfo("Success", "GitLab token saved successfully to token.json")
            self.status_var.set("GitLab token saved")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save token: {str(e)}")
    
    def save_llm_token(self):
        """Save LLM token to file"""
        token = self.llm_token_var.get().strip()
        if not token:
            messagebox.showwarning("Warning", "Please enter an LLM token")
            return
        
        try:
            self.token_manager.save_llm_token(token)
            self.llm_token = token
            messagebox.showinfo("Success", "LLM token saved successfully to llm_token.json")
            self.status_var.set("LLM token saved")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save token: {str(e)}")
    
    def extract_best_practices(self):
        """Extract best practices from checked review comments using Vertafore AI"""
        print("DEBUG: extract_best_practices called")
        
        # Get Vertafore API token
        if not self.llm_token:
            messagebox.showwarning("Warning", "No LLM token found. Please add it to llm_token.json")
            return
        
        print(f"DEBUG: LLM token present: {bool(self.llm_token)}")
        
        # Check if we have comments data
        if not self.comments_data:
            messagebox.showwarning("Warning", "Please fetch comments first")
            return
        
        print(f"DEBUG: comment_checkboxes type: {type(self.comment_checkboxes)}")
        print(f"DEBUG: comment_checkboxes count: {len(self.comment_checkboxes)}")
        
        # Get checked discussions
        checked_discussions = []
        for discussion_id, checkbox_var in self.comment_checkboxes.items():
            if checkbox_var.get():
                # Find the discussion in comments_data
                for discussion in self.comments_data:
                    if discussion.get('id') == discussion_id:
                        checked_discussions.append(discussion)
                        break
        
        print(f"DEBUG: Checked discussions count: {len(checked_discussions)}")
        
        if not checked_discussions:
            messagebox.showwarning("Warning", "Please check at least one discussion in the Comments Review tab")
            return
        
        def extract_in_thread():
            self.progress.start()
            self.status_var.set("Extracting best practices with Vertafore AI...")
            
            try:
                print("DEBUG: Creating LLMService...")
                # Initialize Vertafore LLM service
                llm_service = LLMService(self.llm_token, provider="vertafore")
                
                print("DEBUG: Calling extract_best_practices...")
                # Extract best practices
                success, result = llm_service.extract_best_practices(checked_discussions)
                
                print(f"DEBUG: LLM response - success: {success}")
                
                if success:
                    # Update the best practices tab
                    self.best_practices_text.config(state="normal")
                    self.best_practices_text.delete(1.0, tk.END)
                    
                    # Add header
                    header = f"Extracted Coding Standards from {len(checked_discussions)} Review Discussions\n"
                    header += f"Generated by Claude Sonnet 3.5 via Vertafore Enterprise AI\n"
                    header += "=" * 70 + "\n\n"
                    self.best_practices_text.insert(tk.END, header)
                    
                    # Add LLM response (editable)
                    self.best_practices_text.insert(tk.END, result)
                    
                    # Switch to best practices tab
                    self.notebook.select(self.best_practices_frame)
                    
                    self.status_var.set(f"Successfully extracted coding standards from {len(checked_discussions)} discussions")
                else:
                    print(f"DEBUG: LLM error: {result}")
                    messagebox.showerror("Error", f"Failed to extract best practices: {result}")
                    self.status_var.set("Failed to extract best practices")
                    
            except Exception as e:
                print(f"DEBUG: Exception in extract_in_thread: {str(e)}")
                import traceback
                traceback.print_exc()
                messagebox.showerror("Error", f"Unexpected error: {str(e)}")
                self.status_var.set("Error extracting best practices")
            finally:
                self.progress.stop()
        
        print("DEBUG: Starting extraction thread...")
        threading.Thread(target=extract_in_thread, daemon=True).start()